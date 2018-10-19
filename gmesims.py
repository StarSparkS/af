import string,re,sys,tempfile,shutil,os
from numbers import Number
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter,coordinate_from_string,column_index_from_string

from pprint import pprint as pp
from collections import defaultdict,namedtuple
from itertools import groupby
from functools import reduce

##############################################################################"""
# Helper functions
def create_temporary_copy(path):
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, 'temp_file_name.xlsx')
    shutil.copy2(path, temp_path)
    return temp_path
Tree    = lambda: defaultdict(Tree)
add     = lambda branch,path: reduce(lambda twig,leaf: twig[leaf],path,branch)
# put     = lambda sbranch,path,value: exec(sbranch+'[' + ']['.join(map(repr,path)) + ']'+'='+prep(value))
# put     = lambda sbranch,path,value: exec("%s[%s]=%s" % (sbranch,']['.join(map(repr,path)),prep(value)))
def put(branch,path,value):
    add(branch,path)
    cmd = "branch[%s]=%s" % (']['.join(map(repr,path)),prep(value))
    print(cmd)
    exec(cmd) #,{},dict(branch=branch))
show    = lambda t: {k: show(t[k]) for k in t} if isinstance(t,dict) else t
prep    = lambda x: repr(int(x) if isinstance(x,Number) else x)

Box     = namedtuple('Box','r1 c1 r2 c2'.split())
Cell    = namedtuple('Cell','r c p v'.split())
Line    = namedtuple('Line','rc kv'.split())
pos2rc  = lambda p: (int(p[1]), column_index_from_string(p[0]))
pos2rcs = lambda ps: [pos2rc(p) for p in str(ps).split(':')]
merged  = lambda ws: dict(((r,c),Cell(r,c,get_column_letter(c1)+str(r1),
                                      ws.cell(row=r1,column=c1).value))
                           for (r1,c1),(r2,c2) in map(pos2rcs,ws.merged_cells.ranges)
                           for r in range(r1,r2+1)
                           for c in range(c1,c2+1))

clean = lambda labels,rowmajor=True: dict(
    (k,tuple(filter(None,
                    (list(b)[0].v
                     for a,b in groupby(sorted(labels[k],
                                               key=lambda x: x.r if rowmajor else x.c),
                                        lambda x: x.p)))))
    for k in labels)

def gettags(ws,box,rowmajor=True):
    labels = defaultdict(tuple)
    if box:
        m = merged(ws)
        for r in range(box.r1,box.r2+1):
            for c in range(box.c1,box.c2+1):
                v = ws.cell(row=r,column=c).value
                p = get_column_letter(c)+str(r)
                if v == None:
                    v = m.get((r,c)).v if m.get((r,c)) else None
                    p = m.get((r,c)).p if m.get((r,c)) else None
                labels[c if rowmajor else r] += (Cell(r,c,p,v),)
    return clean(labels,rowmajor)

def unpivot(ws,cbox=False,rbox=False):
    result = []
    clabels = gettags(ws,cbox,True)
    rlabels = gettags(ws,rbox,False)
    if rlabels and clabels:
        for r in range(min(rlabels),max(rlabels)+1):
            for c in range(min(clabels),max(clabels)+1):
                x = ws.cell(row=r,column=c)
                if hasattr(x,'value') and x.value != None:
                    label = rlabels.get(r,()) + clabels.get(c,())
                    result.append(Line(rc=(r,c),kv=dict(label=x.value)))
    elif clabels:
        head = tuple(v for k,v in sorted(clabels.items()))
        min_row = ws.min_row + max(map(len,head))
        min_col = min(clabels)
        max_col = max(clabels)
        for r in range(min_row,ws.max_row+1):
            row = []
            for c in range(min_col,max_col+1):
                x = ws.cell(row=r,column=c)
                row.append(x.value if hasattr(x,'value') else None)
            line = dict(zip(head,row))
            if any(v for v in line.values()):
                result.append(Line(rc=(r,None),kv=line))
    elif rlabels:
        head = tuple(v for k,v in sorted(rlabels.items()))
        min_col = ws.min_column + max(map(len,head))
        min_row = min(rlabels)
        max_row = max(rlabels)
        for c in range(min_col,ws.max_column+1):
            col = []
            for r in range(min_row,max_row+1):
                x = ws.cell(row=r,column=c)
                col.append(x.value if hasattr(x,'value') else None)
            line = dict(zip(head,col))
            if any(v for v in line.values()):
                result.append(Line(rc=(None,c),kv=line))
    else:
        for r in range(ws.min_row,ws.max_row+1):
            row = {}
            for c in range(ws.min_column,ws.max_column+1):
                x = ws.cell(row=r,column=c)
                row[c] = x.value if hasattr(x,'value') else None
            if any(row.items()):
                result.append(Line(rc=(r,None),kv=row))
    return result

###############################################################################
# Configure
print()
wb = load_workbook(create_temporary_copy("gmesIMStest.xlsx"))

config = defaultdict(dict)
for k in [dict((k[0],v) for k,v in kv.items())
          for _,kv in unpivot(wb['Config'],Box(1,1,1,4))]:
    sheet,rc,table,pos = [k[x] for x in 'SHEET RC TABLE POS'.split()]
    rc = 'rbox' if rc == 'R' else 'cbox'
    (r1,c1),(r2,c2) = pos2rcs(pos)
    config[sheet][rc] = Box(r1,c1,r2,c2)
tree = Tree()
for k,v in dict(config).items():
    print(k)
    for rc,kv in [line for line in unpivot(wb[k],**v)]:
        pp((rc,kv))
        for a,b in x.items():
            add(tree[k],a + (b,))
pp(show(tree))
pp(unpivot(wb['Config']))
